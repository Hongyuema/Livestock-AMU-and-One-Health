#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Spearman correlation analyses for livestock antimicrobial use indicators.

The script generates the main and supplementary correlation tables used in the
manuscript. It calculates Spearman's rho, nonparametric bootstrap confidence
intervals, raw p values, and Benjamini-Hochberg false discovery rate q values.
"""

from __future__ import annotations

import argparse
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.stats import ConstantInputWarning, spearmanr


DEFAULT_INPUT_DIR = Path("/Users/mahongyue/Desktop/Processed data")
DEFAULT_OUTPUT_DIR = Path("/Users/mahongyue/Desktop/Results")

GLOBAL_FILE = "抗生素使用效率与总量以及肉类生产消费-校正.xlsx"
DEVELOPMENT_FILE = "抗生素使用效率与总量以及肉类生产消费.xlsx"
CONTINENT_FILE = "大洲分析.xlsx"
OUTPUT_FILE = "All_Tables_Spearman_FDR_bootstrap_recalculated.xlsx"

N_BOOTSTRAP = 10000
RANDOM_SEED = 202406

OUTCOMES = {
    "Total AMU in livestock": "Antimicrobial usage in livestock (tonnes)",
    "PCU-standardized AMU intensity": "Antimicrobial usage in livestock (mg per population corrected units)",
}

MEAT_VARIABLES = [
    ("Total meat production", "Total meat production (tonnes)"),
    ("Pork production", "Pig production (tonnes)"),
    ("Beef and buffalo production", "Beef and buffalo production (tonnes)"),
    ("Sheep and goat production", "Sheep and goat production (tonnes)"),
    ("Poultry production", "Poultry production (tonnes)"),
    ("Total meat consumption", "Total meat consumption"),
    (
        "Pork consumption",
        "pig | 00002733 || Food available for consumption | 0645pc || kilograms per year per capita",
    ),
    (
        "Beef and buffalo consumption",
        "beef and buffalo | 00002731 || Food available for consumption | 0645pc || kilograms per year per capita",
    ),
    (
        "Sheep and goat consumption",
        "sheep and goat | 00002732 || Food available for consumption | 0645pc || kilograms per year per capita",
    ),
    (
        "Poultry consumption",
        "poultry | 00002734 || Food available for consumption | 0645pc || kilograms per year per capita",
    ),
]

DEVELOPMENT_VARIABLES = [
    ("GDP per capita", "GDP per capita"),
    ("Average LAYs (years)", "Learning-adjusted years of schooling"),
]

TABLE_S1_ITEMS = [
    ("Total AMU in livestock", "Antimicrobial usage in livestock (tonnes)"),
    ("PCU-standardized AMU intensity", "Antimicrobial usage in livestock (mg per population corrected units)"),
    ("Total meat production", "Total meat production (tonnes)"),
    ("Pork production", "Pig production (tonnes)"),
    ("Beef and buffalo production", "Beef and buffalo production (tonnes)"),
    ("Sheep and goat production", "Sheep and goat production (tonnes)"),
    ("Poultry production", "Poultry production (tonnes)"),
    ("Total meat consumption", "Total meat consumption"),
    (
        "Pork consumption",
        "pig | 00002733 || Food available for consumption | 0645pc || kilograms per year per capita",
    ),
    (
        "Beef and buffalo consumption",
        "beef and buffalo | 00002731 || Food available for consumption | 0645pc || kilograms per year per capita",
    ),
    (
        "Sheep and goat consumption",
        "sheep and goat | 00002732 || Food available for consumption | 0645pc || kilograms per year per capita",
    ),
    (
        "Poultry consumption",
        "poultry | 00002734 || Food available for consumption | 0645pc || kilograms per year per capita",
    ),
]

TABLE_S3_PAIRS = [
    (
        "Total AMU in livestock",
        "Antimicrobial usage in livestock (tonnes)",
        "Total meat production",
        "Total meat production (tonnes)",
    ),
    (
        "Total AMU in livestock",
        "Antimicrobial usage in livestock (tonnes)",
        "Total meat consumption",
        "Total meat consumption",
    ),
    (
        "PCU-standardized AMU intensity",
        "Antimicrobial usage in livestock (mg per population corrected units)",
        "Total meat production",
        "Total meat production (tonnes)",
    ),
    (
        "PCU-standardized AMU intensity",
        "Antimicrobial usage in livestock (mg per population corrected units)",
        "Total meat consumption",
        "Total meat consumption",
    ),
]

TABLE_S4_ITEMS = [
    ("Total AMU in livestock", "Antimicrobial usage in livestock (tonnes)"),
    ("PCU-standardized AMU intensity", "Antimicrobial usage in livestock (mg per population corrected units)"),
    ("Total meat production", "Total meat production (tonnes)"),
    ("Total meat consumption", "Total meat consumption"),
]

CONTINENT_ORDER = ["Asia", "Europe", "South America", "Africa", "North America", "Oceania"]
REGIONAL_TABLE_ORDER = ["North America", "Oceania", "Africa", "South America", "Europe", "Asia"]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate Spearman correlation tables for livestock AMU analyses."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help="Directory containing the input Excel files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory where the output workbook will be saved.",
    )
    return parser.parse_args()


def read_excel_file(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")
    return pd.read_excel(path)


def available_continents(data: pd.DataFrame, order: list[str]) -> list[str]:
    observed = set(data["Continents"].dropna())
    return [continent for continent in order if continent in observed]


def require_columns(data: pd.DataFrame, columns: list[str], file_name: str) -> None:
    missing = [column for column in columns if column not in data.columns]
    if missing:
        joined = "\n".join(missing)
        raise KeyError(f"Missing columns in {file_name}:\n{joined}")


def validate_input_tables(
    global_df: pd.DataFrame,
    development_df: pd.DataFrame,
    continent_df: pd.DataFrame,
) -> None:
    global_columns = [column for _, column in MEAT_VARIABLES] + list(OUTCOMES.values())
    development_columns = [column for _, column in TABLE_S1_ITEMS]
    development_columns += [column for _, column in DEVELOPMENT_VARIABLES]
    continent_columns = ["Continents"] + list(OUTCOMES.values())
    continent_columns += [column for _, column in TABLE_S4_ITEMS]
    continent_columns += [column for _, column in DEVELOPMENT_VARIABLES]
    continent_columns += [pair[1] for pair in TABLE_S3_PAIRS]
    continent_columns += [pair[3] for pair in TABLE_S3_PAIRS]

    require_columns(global_df, sorted(set(global_columns)), GLOBAL_FILE)
    require_columns(development_df, sorted(set(development_columns)), DEVELOPMENT_FILE)
    require_columns(continent_df, sorted(set(continent_columns)), CONTINENT_FILE)


def log_transform(series: pd.Series) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    return np.log10(values + 1)


def bootstrap_spearman_ci(
    x: np.ndarray,
    y: np.ndarray,
    n_iter: int = N_BOOTSTRAP,
    seed: int = RANDOM_SEED,
) -> tuple[float, float]:
    rng = np.random.default_rng(seed)
    x = np.asarray(x)
    y = np.asarray(y)
    n = len(x)
    estimates = []

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=ConstantInputWarning)
        for _ in range(n_iter):
            idx = rng.integers(0, n, n)
            rho, _ = spearmanr(x[idx], y[idx])
            if np.isfinite(rho):
                estimates.append(rho)

    if not estimates:
        return np.nan, np.nan
    lower, upper = np.percentile(estimates, [2.5, 97.5])
    return float(lower), float(upper)


def bh_fdr(p_values: pd.Series | np.ndarray) -> np.ndarray:
    p_values = np.asarray(p_values, dtype=float)
    n = len(p_values)
    order = np.argsort(p_values)
    ranked = p_values[order]
    adjusted = ranked * n / np.arange(1, n + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    adjusted = np.minimum(adjusted, 1.0)
    q_values = np.empty(n)
    q_values[order] = adjusted
    return q_values


def correlation_record(
    data: pd.DataFrame,
    x_name: str,
    x_col: str,
    y_name: str,
    y_col: str,
    family: str,
    subset_name: str | None = None,
) -> dict[str, object]:
    temp = pd.DataFrame(
        {
            "x": log_transform(data[x_col]),
            "y": log_transform(data[y_col]),
        }
    ).dropna()

    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=ConstantInputWarning)
        rho, p_value = spearmanr(temp["x"], temp["y"])

    ci_low, ci_high = bootstrap_spearman_ci(temp["x"].values, temp["y"].values)

    return {
        "Family": family,
        "Subset": subset_name,
        "X variable": x_name,
        "Y variable": y_name,
        "Spearman's rho": rho,
        "95% CI lower": ci_low,
        "95% CI upper": ci_high,
        "p": p_value,
        "N": len(temp),
    }


def add_fdr(results: pd.DataFrame) -> pd.DataFrame:
    results = results.copy()
    results["FDR q"] = bh_fdr(results["p"].values)
    return results


def format_number(value: float, digits: int = 3) -> str:
    if pd.isna(value):
        return ""
    return f"{value:.{digits}f}".rstrip("0").rstrip(".")


def format_ci(lower: float, upper: float) -> str:
    if pd.isna(lower) or pd.isna(upper):
        return ""
    return f"{lower:.3f} to {upper:.3f}"


def format_p_value(value: float) -> str:
    if pd.isna(value):
        return ""
    if value < 0.001:
        return "<0.001"
    if 0.0495 <= value < 0.051:
        return f"{value:.4f}".rstrip("0").rstrip(".")
    return f"{value:.3f}".rstrip("0").rstrip(".")


def write_matrix(ws, start_row: int, start_col: int, values: list[list[object]]) -> None:
    for row_idx, row in enumerate(values, start=start_row):
        for col_idx, value in enumerate(row, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=value)


def write_dataframe(ws, data: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> None:
    write_matrix(ws, start_row, start_col, [list(data.columns)] + data.values.tolist())


def apply_table_style(ws, header_rows: tuple[int, ...] = (1,), freeze: str | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    thick = Side(style="medium", color="000000")
    thin = Side(style="thin", color="B7B7B7")

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx in header_rows:
        if row_idx <= max_row:
            for cell in ws[row_idx]:
                cell.font = Font(name="Arial", size=10, bold=True)
                cell.fill = header_fill

    if freeze:
        ws.freeze_panes = freeze

    for col_idx in range(1, max_col + 1):
        ws.cell(row=min(header_rows), column=col_idx).border = Border(top=thick)
        ws.cell(row=max(header_rows), column=col_idx).border = Border(bottom=thick)
        ws.cell(row=max_row, column=col_idx).border = Border(bottom=thick)

    for row_idx in range(max(header_rows) + 1, max_row + 1):
        for col_idx in range(1, max_col + 1):
            ws.cell(row=row_idx, column=col_idx).border = Border(bottom=thin)

    for col_idx in range(1, max_col + 1):
        max_len = 0
        for row_idx in range(1, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            max_len = max(max_len, len(str(value)) if value is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 38)


def format_records_for_table(
    results: pd.DataFrame,
    row_items: list[tuple[str, str]],
    y_variables: dict[str, str] | list[tuple[str, str]],
    prefixes: dict[str, str],
) -> pd.DataFrame:
    if isinstance(y_variables, dict):
        y_names = list(y_variables.keys())
    else:
        y_names = [name for name, _ in y_variables]

    rows = []
    for item_name, _ in row_items:
        row = {"Items": item_name}
        for y_name in y_names:
            subset = results[
                (results["X variable"] == item_name) & (results["Y variable"] == y_name)
            ].iloc[0]
            prefix = prefixes[y_name]
            row[f"{prefix} Spearman's rho"] = format_number(subset["Spearman's rho"])
            row[f"{prefix} 95% CI"] = format_ci(subset["95% CI lower"], subset["95% CI upper"])
            row[f"{prefix} p"] = format_p_value(subset["p"])
            row[f"{prefix} FDR q"] = format_p_value(subset["FDR q"])
            row["N"] = int(subset["N"])
        rows.append(row)
    return pd.DataFrame(rows)


def build_table_1(global_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    records = []
    for item_name, predictor_col in MEAT_VARIABLES:
        for outcome_name, outcome_col in OUTCOMES.items():
            records.append(
                correlation_record(
                    global_df,
                    x_name=item_name,
                    x_col=predictor_col,
                    y_name=outcome_name,
                    y_col=outcome_col,
                    family="Table 1",
                )
            )

    results = add_fdr(pd.DataFrame(records))
    table = format_records_for_table(
        results,
        MEAT_VARIABLES,
        OUTCOMES,
        {
            "Total AMU in livestock": "Total AMU",
            "PCU-standardized AMU intensity": "PCU",
        },
    )
    table = table[
        [
            "Items",
            "Total AMU Spearman's rho",
            "Total AMU 95% CI",
            "Total AMU p",
            "Total AMU FDR q",
            "PCU Spearman's rho",
            "PCU 95% CI",
            "PCU p",
            "PCU FDR q",
            "N",
        ]
    ]
    return table, results


def build_table_s1(development_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    records = []
    for item_name, item_col in TABLE_S1_ITEMS:
        for dev_name, dev_col in DEVELOPMENT_VARIABLES:
            records.append(
                correlation_record(
                    development_df,
                    x_name=item_name,
                    x_col=item_col,
                    y_name=dev_name,
                    y_col=dev_col,
                    family="Table S1",
                )
            )

    results = add_fdr(pd.DataFrame(records))
    table = format_records_for_table(
        results,
        TABLE_S1_ITEMS,
        DEVELOPMENT_VARIABLES,
        {"GDP per capita": "GDP", "Average LAYs (years)": "LAYs"},
    )
    table = table[
        [
            "Items",
            "GDP Spearman's rho",
            "GDP 95% CI",
            "GDP p",
            "GDP FDR q",
            "LAYs Spearman's rho",
            "LAYs 95% CI",
            "LAYs p",
            "LAYs FDR q",
            "N",
        ]
    ]
    return table, results


def build_tables_s2_s3(
    continent_df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    continents = available_continents(continent_df, REGIONAL_TABLE_ORDER)
    records_s2 = []
    records_s3 = []

    for continent in continents:
        sub = continent_df[continent_df["Continents"] == continent]
        records_s2.append(
            correlation_record(
                sub,
                x_name="PCU-standardized AMU intensity",
                x_col=OUTCOMES["PCU-standardized AMU intensity"],
                y_name="Total AMU in livestock",
                y_col=OUTCOMES["Total AMU in livestock"],
                family="Table S2+S3",
                subset_name=continent,
            )
        )

        for y_name, y_col, x_name, x_col in TABLE_S3_PAIRS:
            records_s3.append(
                correlation_record(
                    sub,
                    x_name=x_name,
                    x_col=x_col,
                    y_name=y_name,
                    y_col=y_col,
                    family="Table S2+S3",
                    subset_name=continent,
                )
            )

    combined = add_fdr(pd.DataFrame(records_s2 + records_s3))

    s2_results = combined[
        (combined["X variable"] == "PCU-standardized AMU intensity")
        & (combined["Y variable"] == "Total AMU in livestock")
    ].copy()
    s3_results = combined.drop(s2_results.index).copy()

    s2_table = s2_results.rename(columns={"Subset": "Continent"}).copy()
    s2_table["Spearman's rho"] = s2_table["Spearman's rho"].map(format_number)
    s2_table["95% CI"] = s2_results.apply(
        lambda row: format_ci(row["95% CI lower"], row["95% CI upper"]), axis=1
    )
    s2_table["p"] = s2_table["p"].map(format_p_value)
    s2_table["FDR q"] = s2_table["FDR q"].map(format_p_value)
    s2_table = s2_table[["Continent", "Spearman's rho", "95% CI", "p", "FDR q", "N"]]

    rows_s3 = []
    for continent in continents:
        for y_name, _, x_name, _ in TABLE_S3_PAIRS:
            subset = s3_results[
                (s3_results["Subset"] == continent)
                & (s3_results["Y variable"] == y_name)
                & (s3_results["X variable"] == x_name)
            ].iloc[0]
            rows_s3.append(
                {
                    "Continent": continent,
                    "AMU indicator": y_name,
                    "Correlate": x_name,
                    "Spearman's rho": format_number(subset["Spearman's rho"]),
                    "95% CI": format_ci(subset["95% CI lower"], subset["95% CI upper"]),
                    "p": format_p_value(subset["p"]),
                    "FDR q": format_p_value(subset["FDR q"]),
                    "N": int(subset["N"]),
                }
            )
    s3_table = pd.DataFrame(rows_s3)

    combined_table = combined.rename(columns={"Subset": "Continent"}).copy()
    combined_table["Spearman's rho"] = combined_table["Spearman's rho"].map(format_number)
    combined_table["95% CI"] = combined.apply(
        lambda row: format_ci(row["95% CI lower"], row["95% CI upper"]), axis=1
    )
    combined_table["p"] = combined_table["p"].map(format_p_value)
    combined_table["FDR q"] = combined_table["FDR q"].map(format_p_value)
    combined_table = combined_table[
        ["Continent", "Y variable", "X variable", "Spearman's rho", "95% CI", "p", "FDR q", "N"]
    ].rename(columns={"Y variable": "AMU indicator", "X variable": "Correlate"})

    return s2_table, s3_table, combined_table, combined


def build_table_s4(continent_df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    continents = available_continents(continent_df, CONTINENT_ORDER)
    records = []

    for continent in continents:
        sub = continent_df[continent_df["Continents"] == continent]
        for item_name, item_col in TABLE_S4_ITEMS:
            for dev_name, dev_col in DEVELOPMENT_VARIABLES:
                records.append(
                    correlation_record(
                        sub,
                        x_name=item_name,
                        x_col=item_col,
                        y_name=dev_name,
                        y_col=dev_col,
                        family="Table S4",
                        subset_name=continent,
                    )
                )

    results = add_fdr(pd.DataFrame(records))
    rows = []
    for continent in continents:
        for item_name, _ in TABLE_S4_ITEMS:
            row = {"Continent": continent, "Items": item_name}
            for dev_name, _ in DEVELOPMENT_VARIABLES:
                subset = results[
                    (results["Subset"] == continent)
                    & (results["X variable"] == item_name)
                    & (results["Y variable"] == dev_name)
                ].iloc[0]
                prefix = "GDP" if dev_name == "GDP per capita" else "LAYs"
                row[f"{prefix} Spearman's rho"] = format_number(subset["Spearman's rho"])
                row[f"{prefix} 95% CI"] = format_ci(subset["95% CI lower"], subset["95% CI upper"])
                row[f"{prefix} p"] = format_p_value(subset["p"])
                row[f"{prefix} FDR q"] = format_p_value(subset["FDR q"])
                row["N"] = int(subset["N"])
            rows.append(row)

    table = pd.DataFrame(
        rows,
        columns=[
            "Continent",
            "Items",
            "GDP Spearman's rho",
            "GDP 95% CI",
            "GDP p",
            "GDP FDR q",
            "LAYs Spearman's rho",
            "LAYs 95% CI",
            "LAYs p",
            "LAYs FDR q",
            "N",
        ],
    )
    return table, results


def build_table_s5(
    development_df: pd.DataFrame, continent_df: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    records = [
        correlation_record(
            development_df,
            x_name="GDP per capita",
            x_col="GDP per capita",
            y_name="Average LAYs (years)",
            y_col="Learning-adjusted years of schooling",
            family="Table S5",
            subset_name="Globally",
        )
    ]

    for continent in available_continents(continent_df, CONTINENT_ORDER):
        sub = continent_df[continent_df["Continents"] == continent]
        records.append(
            correlation_record(
                sub,
                x_name="GDP per capita",
                x_col="GDP per capita",
                y_name="Average LAYs (years)",
                y_col="Learning-adjusted years of schooling",
                family="Table S5",
                subset_name=continent,
            )
        )

    results = add_fdr(pd.DataFrame(records))
    table = results.rename(columns={"Subset": "Items"}).copy()
    table["Spearman's rho"] = table["Spearman's rho"].map(format_number)
    table["95% CI"] = results.apply(
        lambda row: format_ci(row["95% CI lower"], row["95% CI upper"]), axis=1
    )
    table["p"] = table["p"].map(format_p_value)
    table["FDR q"] = table["FDR q"].map(format_p_value)
    table = table[["Items", "Spearman's rho", "95% CI", "p", "FDR q", "N"]]
    return table, results


def add_formatted_sheet(
    wb: Workbook,
    sheet_name: str,
    table: pd.DataFrame,
    note: str | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    write_dataframe(ws, table)
    apply_table_style(ws, header_rows=(1,), freeze="A2")

    if note:
        note_row = ws.max_row + 2
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=ws.max_column)
        cell = ws.cell(row=note_row, column=1, value=note)
        cell.font = Font(name="Arial", size=9)
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[note_row].height = 70


def add_numeric_sheet(wb: Workbook, sheet_name: str, results: pd.DataFrame) -> None:
    ws = wb.create_sheet(sheet_name)
    write_dataframe(ws, results)
    apply_table_style(ws, header_rows=(1,), freeze="A2")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.000000"


def build_methods_sheet(
    table1_num: pd.DataFrame,
    table_s1_num: pd.DataFrame,
    table_s2s3_num: pd.DataFrame,
    table_s4_num: pd.DataFrame,
    table_s5_num: pd.DataFrame,
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["Transformation", "All continuous variables were transformed using log10(x + 1) before Spearman correlation analysis."],
            ["Correlation", "Spearman's rho was calculated using scipy.stats.spearmanr."],
            ["Confidence intervals", f"Confidence intervals were estimated using nonparametric bootstrap resampling with {N_BOOTSTRAP:,} iterations and RANDOM_SEED = {RANDOM_SEED}."],
            ["Bootstrap procedure", "In each iteration, paired country-level observations were resampled with replacement and Spearman's rho was recalculated."],
            ["Confidence interval limits", "The 2.5th and 97.5th percentiles of the bootstrap distribution were used as the 95% confidence interval."],
            ["Multiple testing", "Raw p values were adjusted using the Benjamini-Hochberg false discovery rate method within each prespecified analysis family."],
            ["Table 1 family", f"{len(table1_num)} tests: 10 production/consumption variables × 2 AMU indicators."],
            ["Table S1 family", f"{len(table_s1_num)} tests: 12 variables × 2 development indicators."],
            ["Table S2+S3 family", f"{len(table_s2s3_num)} tests: continent-stratified regional correlation tests."],
            ["Table S4 family", f"{len(table_s4_num)} continent-stratified development-related correlation tests."],
            ["Table S5 family", f"{len(table_s5_num)} global and continent-stratified GDP-LAYs correlations."],
        ],
        columns=["Item", "Description"],
    )


def write_workbook(
    output_file: Path,
    table1: pd.DataFrame,
    table1_num: pd.DataFrame,
    table_s1: pd.DataFrame,
    table_s1_num: pd.DataFrame,
    table_s2: pd.DataFrame,
    table_s3: pd.DataFrame,
    table_s2s3: pd.DataFrame,
    table_s2s3_num: pd.DataFrame,
    table_s4: pd.DataFrame,
    table_s4_num: pd.DataFrame,
    table_s5: pd.DataFrame,
    table_s5_num: pd.DataFrame,
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    add_formatted_sheet(
        wb,
        "Table 1",
        table1,
        "Notes: All variables were log10(x + 1)-transformed before Spearman correlation analysis. "
        "Confidence intervals were estimated using nonparametric bootstrap resampling with 10,000 iterations. "
        "Raw p values were adjusted using the Benjamini-Hochberg false discovery rate method across the 20 correlations shown in this table.",
    )

    add_formatted_sheet(
        wb,
        "Table S1",
        table_s1,
        "Notes: All variables, including Average LAYs, were log10(x + 1)-transformed before Spearman correlation analysis. "
        f"Raw p values were adjusted using the Benjamini-Hochberg false discovery rate method across the {len(table_s1_num)} development-related correlations shown in this table.",
    )

    add_formatted_sheet(
        wb,
        "Table S2",
        table_s2,
        "Notes: Table S2 and Table S3 were treated as one continent-stratified regional correlation family. "
        f"Raw p values were adjusted together using the Benjamini-Hochberg false discovery rate method across {len(table_s2s3_num)} regional correlation tests.",
    )

    add_formatted_sheet(
        wb,
        "Table S3",
        table_s3,
        "Notes: Table S2 and Table S3 were treated as one continent-stratified regional correlation family. "
        f"Raw p values were adjusted together using the Benjamini-Hochberg false discovery rate method across {len(table_s2s3_num)} regional correlation tests.",
    )

    add_formatted_sheet(
        wb,
        "Table S2+S3 combined",
        table_s2s3,
        "Notes: This table combines the continent-stratified correlations from Tables S2 and S3. "
        f"Raw p values were adjusted using the Benjamini-Hochberg false discovery rate method across all {len(table_s2s3_num)} regional correlation tests.",
    )

    add_formatted_sheet(
        wb,
        "Table S4",
        table_s4,
        "Notes: All variables, including Average LAYs, were log10(x + 1)-transformed before Spearman correlation analysis. "
        f"Raw p values were adjusted using the Benjamini-Hochberg false discovery rate method across the {len(table_s4_num)} continent-stratified development-related correlations shown in this table.",
    )

    add_formatted_sheet(
        wb,
        "Table S5",
        table_s5,
        "Notes: GDP per capita and Average LAYs were log10(x + 1)-transformed before Spearman correlation analysis. "
        f"Raw p values were adjusted using the Benjamini-Hochberg false discovery rate method across the {len(table_s5_num)} GDP-LAYs correlations shown in this table.",
    )

    add_numeric_sheet(wb, "Table1_numeric", table1_num)
    add_numeric_sheet(wb, "TableS1_numeric", table_s1_num)
    add_numeric_sheet(wb, "TableS2S3_numeric", table_s2s3_num)
    add_numeric_sheet(wb, "TableS4_numeric", table_s4_num)
    add_numeric_sheet(wb, "TableS5_numeric", table_s5_num)

    all_numeric = pd.concat(
        [table1_num, table_s1_num, table_s2s3_num, table_s4_num, table_s5_num],
        ignore_index=True,
    )
    add_numeric_sheet(wb, "All_numeric_results", all_numeric)

    methods = build_methods_sheet(table1_num, table_s1_num, table_s2s3_num, table_s4_num, table_s5_num)
    add_formatted_sheet(wb, "Methods", methods)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)


def run(input_dir: Path, output_dir: Path) -> Path:
    global_df = read_excel_file(input_dir / GLOBAL_FILE)
    development_df = read_excel_file(input_dir / DEVELOPMENT_FILE)
    continent_df = read_excel_file(input_dir / CONTINENT_FILE)
    validate_input_tables(global_df, development_df, continent_df)

    table1, table1_num = build_table_1(global_df)
    table_s1, table_s1_num = build_table_s1(development_df)
    table_s2, table_s3, table_s2s3, table_s2s3_num = build_tables_s2_s3(continent_df)
    table_s4, table_s4_num = build_table_s4(continent_df)
    table_s5, table_s5_num = build_table_s5(development_df, continent_df)

    output_file = output_dir / OUTPUT_FILE
    write_workbook(
        output_file,
        table1,
        table1_num,
        table_s1,
        table_s1_num,
        table_s2,
        table_s3,
        table_s2s3,
        table_s2s3_num,
        table_s4,
        table_s4_num,
        table_s5,
        table_s5_num,
    )
    return output_file


def main() -> None:
    args = parse_args()
    output_file = run(args.input_dir, args.output_dir)
    print(f"Saved: {output_file}")


if __name__ == "__main__":
    main()
