#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Reproducible code for country-level livestock antimicrobial-use projection tables.

The script generates the following supplementary tables:
    Table S7: Country-level meat production models and 2050 projections
    Table S8: Bottom-up 2020-anchored change model for livestock AMU in 2050
    Table S9: PCU-standardized intensity-informed production-ratio scaling model

Default input directory:
    /Users/mahongyue/Desktop/Processed data

Default output directory:
    /Users/mahongyue/Desktop/Results
"""

from __future__ import annotations

import argparse
import math
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_INPUT_DIR = Path("/Users/mahongyue/Desktop/Processed data")
DEFAULT_OUTPUT_DIR = Path("/Users/mahongyue/Desktop/Results")

MEAT_FILE = "meat-production-tonnes.csv"
AMU_FILE = "AMU预测.xlsx"
PCU_INTENSITY_FILE = "antibiotic-usage-in-livestock.csv"

REFERENCE_YEAR = 2020
PROJECTION_YEAR = 2050
SCENARIOS = {
    "baseline": 1.0,
    "10% reduction": 0.9,
    "30% reduction": 0.7,
    "50% reduction": 0.5,
}

MEAT_TIME_SERIES_COLUMN = "All meat - Production (tonnes)"
AMU_COLUMN = "Antimicrobial usage in livestock (tonnes)"
MEAT_2020_COLUMN = "Total meat production (tonnes)"
PCU_INTENSITY_COLUMN = "Antimicrobial usage in livestock (mg per population corrected units)"


class MissingColumnError(ValueError):
    pass


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate country-level projection tables for livestock antimicrobial use."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help="Directory containing the input data files.",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory for output tables.",
    )
    parser.add_argument(
        "--meat-file",
        default=MEAT_FILE,
        help="Annual meat-production time-series file.",
    )
    parser.add_argument(
        "--amu-file",
        default=AMU_FILE,
        help="Country-level 2020 AMU and meat-production file.",
    )
    parser.add_argument(
        "--pcu-intensity-file",
        default=PCU_INTENSITY_FILE,
        help="Country-level 2020 PCU-standardized AMU intensity file.",
    )
    return parser.parse_args()


def require_columns(data: pd.DataFrame, columns: list[str], file_name: str) -> None:
    missing = [column for column in columns if column not in data.columns]
    if missing:
        missing_text = "\n".join(missing)
        raise MissingColumnError(f"Missing required columns in {file_name}:\n{missing_text}")


def read_inputs(input_dir: Path, meat_file: str, amu_file: str, pcu_file: str) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    meat_path = input_dir / meat_file
    amu_path = input_dir / amu_file
    pcu_path = input_dir / pcu_file

    for path in [meat_path, amu_path, pcu_path]:
        if not path.exists():
            raise FileNotFoundError(f"Input file not found: {path}")

    meat = pd.read_csv(meat_path)
    amu = pd.read_excel(amu_path)
    pcu = pd.read_csv(pcu_path)

    require_columns(meat, ["Entity", "Code", "Year", MEAT_TIME_SERIES_COLUMN], meat_file)
    require_columns(amu, ["Entity", "Code", AMU_COLUMN, MEAT_2020_COLUMN], amu_file)
    require_columns(pcu, ["Entity", "Code", "Year", PCU_INTENSITY_COLUMN], pcu_file)

    return meat, amu, pcu


def fit_amu_production_slope(amu: pd.DataFrame) -> float:
    model_data = amu[[MEAT_2020_COLUMN, AMU_COLUMN]].dropna().copy()
    fit = sm.OLS(model_data[AMU_COLUMN], sm.add_constant(model_data[[MEAT_2020_COLUMN]])).fit()
    return round(float(fit.params[MEAT_2020_COLUMN]), 10)


def format_p_value(value: float) -> str:
    if pd.isna(value):
        return ""
    if value < 0.001:
        return "<0.001"
    return f"{value:.3f}"


def format_ci_range(lower: float, upper: float) -> str:
    return f"{lower:,.1f} to {upper:,.1f}"


def format_equation_number(value: float) -> str:
    if pd.isna(value):
        return "NA"
    abs_value = abs(value)
    if abs_value >= 1000:
        return f"{value:.1f}"
    return f"{value:.3f}".rstrip("0").rstrip(".")


def fit_country_meat_models(meat: pd.DataFrame, amu: pd.DataFrame) -> pd.DataFrame:
    country_codes = set(amu["Code"].dropna().astype(str))
    meat_country = meat[meat["Code"].astype(str).isin(country_codes)].copy()

    records = []
    for code, group in meat_country.groupby("Code", sort=False):
        g = group.dropna(subset=["Year", MEAT_TIME_SERIES_COLUMN]).copy()
        if len(g) < 3:
            continue

        g["Year centered"] = g["Year"] - REFERENCE_YEAR
        fit = sm.OLS(g[MEAT_TIME_SERIES_COLUMN], sm.add_constant(g[["Year centered"]])).fit()

        alpha = float(fit.params["const"])
        gamma = float(fit.params["Year centered"])
        slope_ci_lower, slope_ci_upper = fit.conf_int().loc["Year centered"].tolist()
        raw_2050 = alpha + gamma * (PROJECTION_YEAR - REFERENCE_YEAR)
        meat_2050 = max(0.0, raw_2050)

        observed_2020 = g.loc[g["Year"] == REFERENCE_YEAR, MEAT_TIME_SERIES_COLUMN]
        if len(observed_2020) > 0 and pd.notna(observed_2020.iloc[0]):
            meat_anchor = float(observed_2020.iloc[0])
            anchor_source = "Observed"
        else:
            meat_anchor = alpha
            anchor_source = "Fitted"

        first_year = int(g["Year"].min())
        last_year = int(g["Year"].max())
        period = f"{first_year}–{last_year}"

        records.append(
            {
                "Entity": g["Entity"].iloc[0],
                "Code": code,
                "n years": int(len(g)),
                "Period": period,
                "Meat 2020 anchor (tonnes)": meat_anchor,
                "2020 anchor source": anchor_source,
                "α, 2020-centered intercept (tonnes)": alpha,
                "γ, annual change (tonnes/year)": gamma,
                "γ 95% CI (tonnes/year)": format_ci_range(float(slope_ci_lower), float(slope_ci_upper)),
                "P for γ": format_p_value(float(fit.pvalues["Year centered"])),
                "R²": float(fit.rsquared),
                "Meat 2050 projection (tonnes)": meat_2050,
                "Zero-constrained 2050 projection": "Yes" if raw_2050 < 0 else "No",
                "Meat-production equation": (
                    f"Meat_i,t = {format_equation_number(alpha)} + "
                    f"{format_equation_number(gamma)} × (Year − 2020)"
                ),
            }
        )

    table = pd.DataFrame(records)
    return table.sort_values(["Entity", "Code"]).reset_index(drop=True)


def build_anchored_change_table(meat_table: pd.DataFrame, amu: pd.DataFrame, beta: float) -> pd.DataFrame:
    source = amu[["Entity", "Code", AMU_COLUMN]].merge(
        meat_table[
            [
                "Code",
                "Meat 2020 anchor (tonnes)",
                "2020 anchor source",
                "Meat 2050 projection (tonnes)",
            ]
        ],
        on="Code",
        how="inner",
    )

    records = []
    for _, row in source.iterrows():
        amu_2020 = float(row[AMU_COLUMN])
        meat_2020 = float(row["Meat 2020 anchor (tonnes)"])
        meat_2050 = float(row["Meat 2050 projection (tonnes)"])
        delta_meat = meat_2050 - meat_2020

        baseline = max(0.0, amu_2020 + SCENARIOS["baseline"] * beta * delta_meat)
        reduction_10 = max(0.0, amu_2020 + SCENARIOS["10% reduction"] * beta * delta_meat)
        reduction_30 = max(0.0, amu_2020 + SCENARIOS["30% reduction"] * beta * delta_meat)
        reduction_50 = max(0.0, amu_2020 + SCENARIOS["50% reduction"] * beta * delta_meat)

        records.append(
            {
                "Entity": row["Entity"],
                "Code": row["Code"],
                "AMU 2020 (tonnes)": amu_2020,
                "Meat 2020 anchor (tonnes)": meat_2020,
                "2020 meat anchor source": row["2020 anchor source"],
                "Meat 2050 projection (tonnes)": meat_2050,
                "ΔMeat 2020–2050 (tonnes)": delta_meat,
                "β (tonnes AMU/tonne meat)": round(beta, 10),
                "AMU 2050 baseline (tonnes)": baseline,
                "AMU 2050 10% reduction (tonnes)": reduction_10,
                "AMU 2050 30% reduction (tonnes)": reduction_30,
                "AMU 2050 50% reduction (tonnes)": reduction_50,
                "Anchored change equation": (
                    f"AMU_i,2050 = {format_equation_number(amu_2020)} + λ × "
                    f"{beta:.7f} × ({format_equation_number(meat_2050)} − "
                    f"{format_equation_number(meat_2020)})"
                ),
            }
        )

    return pd.DataFrame(records).sort_values(["Entity", "Code"]).reset_index(drop=True)


def build_intensity_scaling_table(meat_table: pd.DataFrame, amu: pd.DataFrame, pcu: pd.DataFrame) -> pd.DataFrame:
    pcu_2020 = pcu[pcu["Year"] == REFERENCE_YEAR][["Code", PCU_INTENSITY_COLUMN]].copy()
    source = (
        amu[["Entity", "Code", AMU_COLUMN]]
        .merge(pcu_2020, on="Code", how="left")
        .merge(
            meat_table[
                [
                    "Code",
                    "Meat 2020 anchor (tonnes)",
                    "2020 anchor source",
                    "Meat 2050 projection (tonnes)",
                ]
            ],
            on="Code",
            how="inner",
        )
    )

    records = []
    for _, row in source.iterrows():
        amu_2020 = float(row[AMU_COLUMN])
        meat_2020 = float(row["Meat 2020 anchor (tonnes)"])
        meat_2050 = float(row["Meat 2050 projection (tonnes)"])
        ratio = meat_2050 / meat_2020 if meat_2020 > 0 else np.nan
        constant_intensity = max(0.0, amu_2020 * ratio) if pd.notna(ratio) else np.nan

        records.append(
            {
                "Entity": row["Entity"],
                "Code": row["Code"],
                "AMU 2020 (tonnes)": amu_2020,
                "PCU-standardized AMU intensity 2020 (mg/kg PCU)": row[PCU_INTENSITY_COLUMN],
                "Meat 2020 anchor (tonnes)": meat_2020,
                "2020 meat anchor source": row["2020 anchor source"],
                "Meat 2050 projection (tonnes)": meat_2050,
                "Meat ratio 2050/2020": ratio,
                "AMU 2050 constant 2020 intensity (tonnes)": constant_intensity,
                "AMU 2050 10% reduction (tonnes)": constant_intensity * SCENARIOS["10% reduction"],
                "AMU 2050 30% reduction (tonnes)": constant_intensity * SCENARIOS["30% reduction"],
                "AMU 2050 50% reduction (tonnes)": constant_intensity * SCENARIOS["50% reduction"],
                "Production-ratio scaling equation": (
                    f"AMU_i,2050 = λ × {format_equation_number(amu_2020)} × "
                    f"({format_equation_number(meat_2050)} / {format_equation_number(meat_2020)})"
                ),
            }
        )

    return pd.DataFrame(records).sort_values(["Entity", "Code"]).reset_index(drop=True)


def add_table_sheet(
    workbook: Workbook,
    sheet_name: str,
    title: str,
    notes: list[str],
    data: pd.DataFrame,
) -> None:
    ws = workbook.create_sheet(sheet_name)
    max_col = len(data.columns)

    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = Font(name="Arial", size=11, bold=True)
    ws.cell(1, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)

    for index, note in enumerate(notes, start=2):
        ws.cell(index, 1).value = note
        ws.cell(index, 1).font = Font(name="Arial", size=9)
        ws.cell(index, 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=index, start_column=1, end_row=index, end_column=max_col)

    ws.append([])
    start_row = len(notes) + 3
    for row in dataframe_to_rows(data, index=False, header=True):
        ws.append(row)

    header_row = start_row
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    medium = Side(style="medium", color="7F7F7F")

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=medium, bottom=medium)

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
            if isinstance(cell.value, float):
                if abs(cell.value) >= 1000:
                    cell.number_format = "#,##0.0"
                elif abs(cell.value) >= 1:
                    cell.number_format = "0.000"
                else:
                    cell.number_format = "0.000000"

    text_headers = {
        "Entity",
        "Code",
        "Period",
        "2020 anchor source",
        "2020 meat anchor source",
        "Zero-constrained 2050 projection",
        "Meat-production equation",
        "Anchored change equation",
        "Production-ratio scaling equation",
    }
    for column_index, header_cell in enumerate(ws[header_row], start=1):
        if header_cell.value in text_headers:
            for cell in ws.iter_rows(
                min_row=header_row + 1,
                max_row=ws.max_row,
                min_col=column_index,
                max_col=column_index,
            ):
                cell[0].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for column_index in range(1, max_col + 1):
        column_letter = get_column_letter(column_index)
        values = [ws.cell(row, column_index).value for row in range(1, min(ws.max_row, 200) + 1)]
        width = max(len(str(value)) for value in values if value is not None) + 2
        ws.column_dimensions[column_letter].width = min(max(width, 10), 42)

    ws.freeze_panes = f"A{header_row + 1}"

    table_ref = f"A{header_row}:{get_column_letter(max_col)}{ws.max_row}"
    table_name = "Table_" + "".join(ch for ch in sheet_name if ch.isalnum())[:20]
    tab = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    tab.tableStyleInfo = style
    ws.add_table(tab)


def write_outputs(output_dir: Path, table_s7: pd.DataFrame, table_s8: pd.DataFrame, table_s9: pd.DataFrame, beta: float) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    table_s7_csv = output_dir / "TableS7_country_meat_production_projection.csv"
    table_s8_csv = output_dir / "TableS8_2020_anchored_change_AMU_projection.csv"
    table_s9_csv = output_dir / "TableS9_PCU_intensity_production_scaling_AMU_projection.csv"
    workbook_path = output_dir / "TableS7_S8_S9_country_anchored_projection_tables.xlsx"
    zip_path = output_dir / "TableS7_S8_S9_country_projection_package.zip"

    table_s7.to_csv(table_s7_csv, index=False)
    table_s8.to_csv(table_s8_csv, index=False)
    table_s9.to_csv(table_s9_csv, index=False)

    wb = Workbook()
    wb.remove(wb.active)

    add_table_sheet(
        wb,
        "Table S7",
        "Table S7. Country-level meat production models and 2050 projections",
        [
            "Model: Meat_i,t = α_i + γ_i × (Year − 2020), fitted separately for each country using annual meat production time-series. 2050 projections below zero were constrained to zero.",
            "Rows include countries matched to 2020 AMU data. When observed 2020 meat production was unavailable in the time-series, the fitted 2020 intercept was used as the meat-production anchor.",
        ],
        table_s7,
    )

    baseline_sum = table_s8["AMU 2050 baseline (tonnes)"].sum() / 1000
    add_table_sheet(
        wb,
        "Table S8",
        "Table S8. Bottom-up 2020-anchored change model for livestock AMU in 2050",
        [
            f"Model: AMU_i,2050 = AMU_i,2020 + λ × β × (Meat_i,2050 − Meat_i,2020), where β = {beta:.7f} tonnes AMU per tonne meat production and λ = 1.0, 0.9, 0.7, or 0.5.",
            f"Global baseline 2050 estimate from this harmonized table: {baseline_sum:.1f} thousand tonnes. This is a bottom-up sensitivity analysis, not a country-specific forecast.",
        ],
        table_s8,
    )

    scaling_sum = table_s9["AMU 2050 constant 2020 intensity (tonnes)"].sum() / 1000
    add_table_sheet(
        wb,
        "Table S9",
        "Table S9. PCU-standardized intensity-informed production-ratio scaling for livestock AMU in 2050",
        [
            "Model: AMU_i,2050 = AMU_i,2020 × (Meat_i,2050 / Meat_i,2020) × λ. The 2020 PCU-standardized intensity column is retained as baseline intensity information; future PCU is not independently forecast.",
            f"Global constant-scaling 2050 estimate: {scaling_sum:.1f} thousand tonnes. The PCU-proxy formulation is algebraically equivalent to production-ratio scaling from observed 2020 AMU.",
        ],
        table_s9,
    )

    wb.save(workbook_path)

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in [workbook_path, table_s7_csv, table_s8_csv, table_s9_csv, Path(__file__)]:
            if path.exists():
                zf.write(path, arcname=path.name)

    print("Saved:")
    print(workbook_path)
    print(table_s7_csv)
    print(table_s8_csv)
    print(table_s9_csv)
    print(zip_path)


def main() -> None:
    args = parse_args()
    meat, amu, pcu = read_inputs(args.input_dir, args.meat_file, args.amu_file, args.pcu_intensity_file)

    beta = fit_amu_production_slope(amu)
    table_s7 = fit_country_meat_models(meat, amu)
    table_s8 = build_anchored_change_table(table_s7, amu, beta)
    table_s9 = build_intensity_scaling_table(table_s7, amu, pcu)

    write_outputs(args.output_dir, table_s7, table_s8, table_s9, beta)


if __name__ == "__main__":
    main()
