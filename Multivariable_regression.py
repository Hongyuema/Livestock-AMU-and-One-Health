#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Multivariable robustness analysis for country-level livestock antimicrobial-use indicators.

This script generates Supplementary Table S6 and a supplementary forest plot from the
country-level analysis dataset. Continuous variables are transformed using log10(x + 1),
standardized within each complete-case model sample, and analysed using ordinary least-
squares regression with HC3 robust standard errors.
"""

from __future__ import annotations

import argparse
import textwrap
import zipfile
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from statsmodels.stats.multitest import multipletests
from statsmodels.stats.outliers_influence import variance_inflation_factor


DEFAULT_INPUT_DIR = Path("/Users/mahongyue/Desktop/Processed data")
DEFAULT_OUTPUT_DIR = Path("/Users/mahongyue/Desktop/Results")
DATA_FILE_NAME = "大洲分析.xlsx"

TABLE_OUTPUT = "TableS6_multivariable_robustness_models_iScience_final.xlsx"
FOREST_PNG_OUTPUT = "FigureS_multivariable_robustness_forest_iScience_final.png"
FOREST_SVG_OUTPUT = "FigureS_multivariable_robustness_forest_iScience_final.svg"
FOREST_PDF_OUTPUT = "FigureS_multivariable_robustness_forest_iScience_final.pdf"
TEXT_OUTPUT = "TableS6_multivariable_methods_and_results.txt"
ZIP_OUTPUT = "TableS6_multivariable_robustness_package.zip"


COLUMNS = {
    "Total AMU in livestock": "Antimicrobial usage in livestock (tonnes)",
    "PCU-standardized AMU intensity": "Antimicrobial usage in livestock (mg per population corrected units)",
    "Total meat production": "Total meat production (tonnes)",
    "Total meat consumption": "Total meat consumption",
    "GDP per capita": "GDP per capita",
    "Average LAYs": "Learning-adjusted years of schooling",
}

OUTCOMES = [
    "Total AMU in livestock",
    "PCU-standardized AMU intensity",
]

DISPLAY_PREDICTORS = [
    "Total meat production",
    "Total meat consumption",
    "GDP per capita",
    "Average LAYs",
]

MODELS = [
    {
        "Model ID": "M1",
        "Model": "Development-adjusted model",
        "Continent fixed effects": "No",
        "include_continent": False,
    },
    {
        "Model ID": "M2",
        "Model": "Region-adjusted sensitivity model",
        "Continent fixed effects": "Yes",
        "include_continent": True,
    },
]


class AnalysisPaths:
    def __init__(self, input_dir: Path, output_dir: Path) -> None:
        self.input_dir = input_dir
        self.output_dir = output_dir
        self.data_file = input_dir / DATA_FILE_NAME
        self.table = output_dir / TABLE_OUTPUT
        self.forest_png = output_dir / FOREST_PNG_OUTPUT
        self.forest_svg = output_dir / FOREST_SVG_OUTPUT
        self.forest_pdf = output_dir / FOREST_PDF_OUTPUT
        self.text = output_dir / TEXT_OUTPUT
        self.archive = output_dir / ZIP_OUTPUT


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate Table S6 multivariable robustness models and the supplementary forest plot."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help="Directory containing the input Excel file. Default: %(default)s",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help="Directory where output files will be written. Default: %(default)s",
    )
    return parser.parse_args()


def require_input(paths: AnalysisPaths) -> None:
    if not paths.data_file.exists():
        raise FileNotFoundError(f"Input file not found: {paths.data_file}")
    paths.output_dir.mkdir(parents=True, exist_ok=True)


def require_columns(raw: pd.DataFrame) -> None:
    required = set(COLUMNS.values()) | {"Continents", "Entity", "Code"}
    missing = sorted(column for column in required if column not in raw.columns)
    if missing:
        missing_text = "\n".join(missing)
        raise KeyError(f"The following required columns are missing from {DATA_FILE_NAME}:\n{missing_text}")


def log_transform(values: pd.Series) -> pd.Series:
    return np.log10(pd.to_numeric(values, errors="coerce") + 1)


def z_score(values: pd.Series) -> pd.Series:
    standard_deviation = values.std(ddof=0)
    if standard_deviation == 0 or pd.isna(standard_deviation):
        raise ValueError(f"Cannot standardize a constant or empty variable: {values.name}")
    return (values - values.mean()) / standard_deviation


def format_float(value: float, digits: int = 3) -> str:
    if pd.isna(value):
        return ""
    return f"{value:.{digits}f}"


def format_p(value: float) -> str:
    if pd.isna(value):
        return ""
    if value < 0.001:
        return "<0.001"
    if 0.0495 <= value < 0.051:
        return f"{value:.4f}"
    return f"{value:.3f}"


def format_ci(lower: float, upper: float) -> str:
    return f"{lower:.3f} to {upper:.3f}"


def prepare_data(raw: pd.DataFrame) -> pd.DataFrame:
    data = pd.DataFrame({name: log_transform(raw[column]) for name, column in COLUMNS.items()})
    data["Continent"] = raw["Continents"]
    data["Entity"] = raw["Entity"]
    data["Code"] = raw["Code"]
    return data


def calculate_vif(design_matrix: pd.DataFrame) -> dict[str, float]:
    vif_values = {}
    for index, column in enumerate(design_matrix.columns):
        if column == "const":
            continue
        vif_values[column] = variance_inflation_factor(design_matrix.values, index)
    return vif_values


def fit_model(data: pd.DataFrame, outcome: str, model_spec: dict) -> tuple[list[dict], dict]:
    required_columns = [outcome] + DISPLAY_PREDICTORS
    if model_spec["include_continent"]:
        required_columns.append("Continent")

    model_data = data[required_columns].dropna().copy()

    for column in [outcome] + DISPLAY_PREDICTORS:
        model_data[column] = z_score(model_data[column])

    predictors = model_data[DISPLAY_PREDICTORS].copy()
    if model_spec["include_continent"]:
        continent_terms = pd.get_dummies(
            model_data["Continent"],
            prefix="Continent",
            drop_first=True,
        ).astype(float)
        predictors = pd.concat([predictors, continent_terms], axis=1)

    design_matrix = sm.add_constant(predictors, has_constant="add")
    response = model_data[outcome]
    fitted_model = sm.OLS(response, design_matrix).fit(cov_type="HC3")

    vif_values = calculate_vif(design_matrix)
    displayed_p_values = [fitted_model.pvalues[predictor] for predictor in DISPLAY_PREDICTORS]
    displayed_q_values = multipletests(displayed_p_values, method="fdr_bh")[1]
    q_value_by_predictor = dict(zip(DISPLAY_PREDICTORS, displayed_q_values))
    confidence_intervals = fitted_model.conf_int()

    rows = []
    for predictor in DISPLAY_PREDICTORS:
        rows.append(
            {
                "Outcome": outcome,
                "Model ID": model_spec["Model ID"],
                "Model": model_spec["Model"],
                "Predictor": predictor,
                "Standardized beta": fitted_model.params[predictor],
                "95% CI lower": confidence_intervals.loc[predictor, 0],
                "95% CI upper": confidence_intervals.loc[predictor, 1],
                "p value": fitted_model.pvalues[predictor],
                "FDR q": q_value_by_predictor[predictor],
                "VIF": vif_values[predictor],
                "N": int(fitted_model.nobs),
                "Adjusted R2": fitted_model.rsquared_adj,
                "Continent fixed effects": model_spec["Continent fixed effects"],
            }
        )

    summary = {
        "Outcome": outcome,
        "Model ID": model_spec["Model ID"],
        "Model": model_spec["Model"],
        "N": int(fitted_model.nobs),
        "Adjusted R2": fitted_model.rsquared_adj,
        "R2": fitted_model.rsquared,
        "AIC": fitted_model.aic,
        "BIC": fitted_model.bic,
        "Continent fixed effects": model_spec["Continent fixed effects"],
        "Displayed predictors max VIF": max(vif_values[predictor] for predictor in DISPLAY_PREDICTORS),
    }

    return rows, summary


def run_analysis(paths: AnalysisPaths) -> tuple[pd.DataFrame, pd.DataFrame]:
    raw = pd.read_excel(paths.data_file)
    require_columns(raw)
    data = prepare_data(raw)

    rows = []
    summaries = []
    for outcome in OUTCOMES:
        for model_spec in MODELS:
            model_rows, model_summary = fit_model(data, outcome, model_spec)
            rows.extend(model_rows)
            summaries.append(model_summary)

    return pd.DataFrame(rows), pd.DataFrame(summaries)


def make_display_table(results: pd.DataFrame) -> pd.DataFrame:
    table = results.copy()
    table["Standardized β"] = table["Standardized beta"].apply(lambda value: format_float(value, 3))
    table["95% CI"] = table.apply(lambda row: format_ci(row["95% CI lower"], row["95% CI upper"]), axis=1)
    table["p value"] = table["p value"].apply(format_p)
    table["FDR q"] = table["FDR q"].apply(format_p)
    table["VIF"] = table["VIF"].apply(lambda value: format_float(value, 2))
    table["Adjusted R²"] = table["Adjusted R2"].apply(lambda value: format_float(value, 3))
    return table[
        [
            "Outcome",
            "Model ID",
            "Model",
            "Predictor",
            "Standardized β",
            "95% CI",
            "p value",
            "FDR q",
            "VIF",
            "N",
            "Adjusted R²",
            "Continent fixed effects",
        ]
    ]


def apply_header_style(cell) -> None:
    cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_cell_style(cell) -> None:
    cell.font = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_table_sheet(workbook: Workbook, display_table: pd.DataFrame) -> None:
    worksheet = workbook.active
    worksheet.title = "Table S6"
    worksheet["A1"] = "Table S6. Multivariable robustness models for country-level livestock antimicrobial use"
    worksheet["A1"].font = Font(name="Arial", size=12, bold=True)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(display_table.columns))

    header_row = 3
    for column_index, column_name in enumerate(display_table.columns, start=1):
        cell = worksheet.cell(row=header_row, column=column_index, value=column_name)
        apply_header_style(cell)

    for row_index, row in enumerate(display_table.itertuples(index=False), start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            apply_cell_style(cell)

    for row_index in range(header_row + 1, header_row + 1 + len(display_table)):
        model_id = worksheet.cell(row=row_index, column=2).value
        fill = PatternFill("solid", fgColor="F8FBF7" if model_id == "M2" else "FFFFFF")
        for column_index in range(1, len(display_table.columns) + 1):
            worksheet.cell(row=row_index, column=column_index).fill = fill

    q_column = list(display_table.columns).index("FDR q") + 1
    for row_index in range(header_row + 1, header_row + 1 + len(display_table)):
        q_value = worksheet.cell(row=row_index, column=q_column).value
        significant = q_value == "<0.001"
        if not significant:
            try:
                significant = float(q_value) < 0.05
            except (TypeError, ValueError):
                significant = False
        if significant:
            for column_index in [4, 5, 6, 7, 8]:
                worksheet.cell(row=row_index, column=column_index).font = Font(name="Arial", size=10, bold=True)

    thin = Side(style="thin", color="D9E2EC")
    thick = Side(style="medium", color="1F4E79")
    max_row = worksheet.max_row
    max_col = worksheet.max_column
    for row_index in range(header_row, max_row + 1):
        for column_index in range(1, max_col + 1):
            worksheet.cell(row=row_index, column=column_index).border = Border(bottom=thin)
    for column_index in range(1, max_col + 1):
        worksheet.cell(row=header_row, column=column_index).border = Border(top=thick, bottom=thick)
        worksheet.cell(row=max_row, column=column_index).border = Border(bottom=thick)

    column_widths = {
        "A": 30,
        "B": 10,
        "C": 32,
        "D": 26,
        "E": 15,
        "F": 18,
        "G": 12,
        "H": 12,
        "I": 10,
        "J": 8,
        "K": 12,
        "L": 18,
    }
    for column_letter, width in column_widths.items():
        worksheet.column_dimensions[column_letter].width = width
    worksheet.row_dimensions[1].height = 24
    worksheet.row_dimensions[header_row].height = 32
    worksheet.freeze_panes = "A4"

    note_row = max_row + 2
    note = (
        "Notes: Total AMU in livestock and PCU-standardized AMU intensity were analysed as separate outcomes. "
        "All continuous variables were transformed using log10(x + 1) and then standardized to z-scores within each model-specific complete-case sample. "
        "Coefficients are standardized beta estimates from ordinary least-squares linear regression with HC3 robust standard errors. "
        "M1 is the development-adjusted model including total meat production, total meat consumption, GDP per capita and Average LAYs. "
        "M2 is the region-adjusted sensitivity model additionally including continent fixed effects. "
        "Continent fixed effects were included as adjustment covariates and are not shown. "
        "FDR q values were calculated using the Benjamini-Hochberg method across the four displayed coefficients within each fitted model. "
        "VIF, variance inflation factor; no imputation was performed."
    )
    worksheet.cell(row=note_row, column=1, value=note)
    worksheet.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=max_col)
    worksheet.cell(row=note_row, column=1).font = Font(name="Arial", size=9)
    worksheet.cell(row=note_row, column=1).alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    worksheet.row_dimensions[note_row].height = 80


def write_dataframe_sheet(workbook: Workbook, sheet_name: str, data: pd.DataFrame, numeric: bool = False) -> None:
    worksheet = workbook.create_sheet(sheet_name)
    for column_index, column_name in enumerate(data.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        apply_header_style(cell)

    for row_index, row in enumerate(data.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, np.floating):
                value = float(value)
            cell = worksheet.cell(row=row_index, column=column_index, value=value)
            apply_cell_style(cell)
            if numeric and isinstance(value, float):
                cell.number_format = "0.000000"

    for column_index in range(1, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 22 if sheet_name != "Methods note" else 36
    worksheet.freeze_panes = "A2"


def make_method_note() -> pd.DataFrame:
    return pd.DataFrame(
        [
            ["Analysis purpose", "Focused multivariable robustness analysis for the two livestock antimicrobial-use indicators."],
            ["Outcomes", "Total AMU in livestock; PCU-standardized AMU intensity."],
            ["Transformation", "All continuous variables were transformed using log10(x + 1)."],
            ["Standardization", "Continuous outcomes and predictors were standardized to z-scores within each model-specific complete-case sample."],
            ["Model M1", "Development-adjusted model: outcome ~ total meat production + total meat consumption + GDP per capita + Average LAYs."],
            ["Model M2", "Region-adjusted sensitivity model: M1 covariates + continent fixed effects."],
            ["Displayed coefficients", "Only the four prespecified continuous predictors are displayed; continent fixed-effect coefficients are adjustment covariates and are not shown."],
            ["Standard errors", "HC3 robust standard errors."],
            ["Multiple testing", "Benjamini-Hochberg false discovery rate correction across the four displayed coefficients within each fitted model."],
            ["Missing values", "Model-specific complete-case analysis; no imputation was performed."],
            ["Multicollinearity", "Variance inflation factors were calculated from each fitted model design matrix."],
        ],
        columns=["Item", "Description"],
    )


def write_excel(results: pd.DataFrame, summaries: pd.DataFrame, paths: AnalysisPaths) -> None:
    display_table = make_display_table(results)
    numeric_results = results[
        [
            "Outcome",
            "Model ID",
            "Model",
            "Predictor",
            "Standardized beta",
            "95% CI lower",
            "95% CI upper",
            "p value",
            "FDR q",
            "VIF",
            "N",
            "Adjusted R2",
            "Continent fixed effects",
        ]
    ].copy()

    summary_display = summaries.copy()
    for column in ["Adjusted R2", "R2"]:
        summary_display[column] = summary_display[column].map(lambda value: round(value, 6))
    for column in ["AIC", "BIC", "Displayed predictors max VIF"]:
        summary_display[column] = summary_display[column].map(lambda value: round(value, 3))

    workbook = Workbook()
    write_table_sheet(workbook, display_table)
    write_dataframe_sheet(workbook, "Model summary", summary_display, numeric=True)
    write_dataframe_sheet(workbook, "Numeric results", numeric_results, numeric=True)
    write_dataframe_sheet(workbook, "Methods note", make_method_note())
    workbook.save(paths.table)


def make_forest_plot(results: pd.DataFrame, paths: AnalysisPaths) -> None:
    plot_data = results[results["Model ID"] == "M2"].copy()

    labels = []
    y_positions = []
    betas = []
    lower_bounds = []
    upper_bounds = []
    q_values = []
    y_position = 0

    for outcome in OUTCOMES:
        labels.append(outcome)
        y_positions.append(y_position)
        betas.append(np.nan)
        lower_bounds.append(np.nan)
        upper_bounds.append(np.nan)
        q_values.append("")
        y_position += 1

        subset = plot_data[plot_data["Outcome"] == outcome]
        for predictor in DISPLAY_PREDICTORS[::-1]:
            row = subset[subset["Predictor"] == predictor].iloc[0]
            labels.append("   " + predictor)
            y_positions.append(y_position)
            betas.append(row["Standardized beta"])
            lower_bounds.append(row["95% CI lower"])
            upper_bounds.append(row["95% CI upper"])
            q_values.append(format_p(row["FDR q"]))
            y_position += 1
        y_position += 1

    finite_positions = [index for index, beta in enumerate(betas) if np.isfinite(beta)]
    finite_betas = np.array([betas[index] for index in finite_positions])
    finite_y = np.array([y_positions[index] for index in finite_positions])
    finite_lower = np.array([lower_bounds[index] for index in finite_positions])
    finite_upper = np.array([upper_bounds[index] for index in finite_positions])

    fig, ax = plt.subplots(figsize=(8.2, 5.2))
    ax.errorbar(
        finite_betas,
        finite_y,
        xerr=[finite_betas - finite_lower, finite_upper - finite_betas],
        fmt="o",
        capsize=3,
        linewidth=1.5,
        markersize=5,
    )
    ax.axvline(0, linestyle="--", linewidth=1)
    ax.set_yticks(y_positions)
    ax.set_yticklabels(labels, fontsize=9)
    ax.set_xlabel("Standardized beta (95% CI)", fontsize=10)
    ax.set_title("Region-adjusted multivariable robustness model", fontsize=12, weight="bold", pad=12)
    ax.set_xlim(-0.75, 1.15)
    ax.set_ylim(-0.8, max(y_positions) + 0.8)
    ax.invert_yaxis()

    for tick_label, label in zip(ax.get_yticklabels(), labels):
        if not label.startswith("   "):
            tick_label.set_fontweight("bold")

    x_text = 1.03
    for position, beta, q_value in zip(y_positions, betas, q_values):
        if np.isfinite(beta):
            q_label = f"q{q_value}" if str(q_value).startswith("<") else f"q={q_value}"
            ax.text(x_text, position, q_label, va="center", fontsize=8)
    ax.text(x_text, -0.5, "FDR", va="center", fontsize=8, weight="bold")

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(axis="x", linestyle=":", linewidth=0.5)
    fig.tight_layout()
    fig.savefig(paths.forest_png, dpi=600, bbox_inches="tight")
    fig.savefig(paths.forest_svg, bbox_inches="tight")
    fig.savefig(paths.forest_pdf, bbox_inches="tight")
    plt.close(fig)


def get_result_row(results: pd.DataFrame, outcome: str, model_id: str, predictor: str) -> pd.Series:
    return results[
        (results["Outcome"] == outcome)
        & (results["Model ID"] == model_id)
        & (results["Predictor"] == predictor)
    ].iloc[0]


def write_text_summary(results: pd.DataFrame, paths: AnalysisPaths) -> None:
    total_m1 = get_result_row(results, "Total AMU in livestock", "M1", "Total meat production")
    total_m2 = get_result_row(results, "Total AMU in livestock", "M2", "Total meat production")
    consumption_m2 = get_result_row(results, "Total AMU in livestock", "M2", "Total meat consumption")
    pcu_production_m2 = get_result_row(results, "PCU-standardized AMU intensity", "M2", "Total meat production")
    pcu_consumption_m2 = get_result_row(results, "PCU-standardized AMU intensity", "M2", "Total meat consumption")

    text = f"""
Methods text

Multivariable linear regression models were fitted as focused robustness analyses to evaluate whether the main country-level antimicrobial-use associations remained evident after adjustment for development-related and regional covariates. Total AMU in livestock and PCU-standardized AMU intensity were analysed as separate outcomes. Antimicrobial-use indicators, meat production, meat consumption, GDP per capita, and Average LAYs were transformed using log10(x + 1), and continuous outcomes and predictors were then standardized to z-scores within each model-specific complete-case sample. Two prespecified models were fitted for each outcome: a development-adjusted model including total meat production, total meat consumption, GDP per capita, and Average LAYs, and a region-adjusted sensitivity model additionally including continent fixed effects. HC3 robust standard errors were used. Multicollinearity was assessed using variance inflation factors. Missing values were handled using model-specific complete-case analysis, and no imputation was performed. Raw p values for the four displayed coefficients within each fitted model were adjusted using the Benjamini-Hochberg false discovery rate method.

Results text

To assess whether the main country-level associations were robust to adjustment for development-related and regional covariates, prespecified multivariable linear regression models were fitted for total AMU in livestock and PCU-standardized AMU intensity. Total meat production remained strongly associated with total AMU after adjustment for total meat consumption, GDP per capita, and Average LAYs (standardized beta = {total_m1['Standardized beta']:.3f}, 95% CI {total_m1['95% CI lower']:.3f} to {total_m1['95% CI upper']:.3f}, FDR q {format_p(total_m1['FDR q'])}). This association was retained in the region-adjusted sensitivity model including continent fixed effects (standardized beta = {total_m2['Standardized beta']:.3f}, 95% CI {total_m2['95% CI lower']:.3f} to {total_m2['95% CI upper']:.3f}, FDR q {format_p(total_m2['FDR q'])}), whereas total meat consumption was not independently associated with total AMU (standardized beta = {consumption_m2['Standardized beta']:.3f}, 95% CI {consumption_m2['95% CI lower']:.3f} to {consumption_m2['95% CI upper']:.3f}, FDR q {format_p(consumption_m2['FDR q'])}). For PCU-standardized AMU intensity, adjusted associations with total meat production (standardized beta = {pcu_production_m2['Standardized beta']:.3f}, 95% CI {pcu_production_m2['95% CI lower']:.3f} to {pcu_production_m2['95% CI upper']:.3f}, FDR q {format_p(pcu_production_m2['FDR q'])}) and total meat consumption (standardized beta = {pcu_consumption_m2['Standardized beta']:.3f}, 95% CI {pcu_consumption_m2['95% CI lower']:.3f} to {pcu_consumption_m2['95% CI upper']:.3f}, FDR q {format_p(pcu_consumption_m2['FDR q'])}) were weaker after regional adjustment. These models support the interpretation that absolute AMU burden and PCU-standardized AMU intensity capture distinct dimensions of livestock antimicrobial pressure.

Supplementary figure legend

Figure Sx. Region-adjusted multivariable robustness models for country-level livestock antimicrobial use. Forest plot showing standardized beta estimates and 95% confidence intervals from region-adjusted sensitivity models. Total AMU in livestock and PCU-standardized AMU intensity were analysed as separate outcomes. All continuous variables were log10(x + 1)-transformed and standardized before modelling. Models included total meat production, total meat consumption, GDP per capita, Average LAYs, and continent fixed effects. Continent fixed-effect coefficients were included as adjustment covariates and are not shown.
"""
    paths.text.write_text(textwrap.dedent(text).strip() + "\n", encoding="utf-8")


def create_archive(paths: AnalysisPaths) -> None:
    files_to_archive = [
        paths.table,
        paths.forest_png,
        paths.forest_svg,
        paths.forest_pdf,
        paths.text,
        Path(__file__),
    ]
    with zipfile.ZipFile(paths.archive, "w", zipfile.ZIP_DEFLATED) as archive:
        for file_path in files_to_archive:
            if file_path.exists():
                archive.write(file_path, arcname=file_path.name)


def main() -> None:
    args = parse_args()
    paths = AnalysisPaths(args.input_dir, args.output_dir)
    require_input(paths)

    results, summaries = run_analysis(paths)
    write_excel(results, summaries, paths)
    make_forest_plot(results, paths)
    write_text_summary(results, paths)
    create_archive(paths)

    print("Saved output files:")
    print(paths.table)
    print(paths.forest_png)
    print(paths.forest_svg)
    print(paths.forest_pdf)
    print(paths.text)
    print(paths.archive)


if __name__ == "__main__":
    main()
